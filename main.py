from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

priceMin = '200'
priceMax = '700'
chrome_options = Options()
chrome_options.add_argument("--disable-search-engine-choice-screen")
driver = webdriver.Chrome(options=chrome_options)

driver.get(
    f"https://www.alza.sk/alza-dni/e39.htm#f&limit={priceMin}--{priceMax}&cst=0,2&cud=0&pg=1-300")  # Load the initial page

reject_cookies_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.CSS_SELECTOR, "a.cookies-info__button--link.js-cookies-info-reject"))
)
reject_cookies_button.click()

WebDriverWait(driver, 240).until(
    EC.invisibility_of_element((By.CLASS_NAME, "circle-loader-container"))
)

page_source = driver.page_source

# Parse the page source with BeautifulSoup
soup = BeautifulSoup(page_source, 'html.parser')

# Find all spans that match the criteria (containing "25", "30", etc.)
target_discounts = ["20"]  # Adjust this list as needed
discount_items = []

for discount in target_discounts:
    spans = soup.find_all("span", class_="coupon-block__label--code", string=lambda t: t and discount in t)
    for span in spans:
        parent_div = span.find_parent("div", class_="box browsingitem js-box canBuy inStockAvailability")
        if parent_div:
            product_link = parent_div.find("a", class_="name browsinglink js-box-link")
            price_before = parent_div.find("span", class_="price-box__price")
            price_after = parent_div.find("span", class_="coupon-block__price")
            avl = parent_div.find("span", class_="avlVal avl2 none")

            if product_link:
                discount_items.append({
                    'product_name': product_link.text.strip(),
                    'product_url': f"https://www.alza.sk{product_link['href']}",
                    'discount_percent': discount,
                    'price_before': price_before.text.strip() if price_before else 'N/A',
                    'price_after': price_after.text.strip() if price_after else 'N/A'
                })

discount_items.sort(
    key=lambda i: (-float(i['discount_percent']), -float(i['price_after'].strip('€').replace(',', '.'))))
# Output the results
# Convert the list to a DataFrame
df = pd.DataFrame(discount_items)

# Create the dynamic filename
output_file = f'#{'#'.join(map(str, target_discounts))}_percent-discounted-items_{datetime.now().strftime('%d-%m-%Y@%H-%M-%S')}.xlsx'
df.to_excel(output_file, index=False)

# Close the WebDriver
driver.quit()

wb = load_workbook(output_file)
ws = wb.active

# Adjust column widths
ws.column_dimensions[get_column_letter(1)].width = 60  # product_name column
ws.column_dimensions[get_column_letter(2)].width = 40  # product_url column

# Iterate over the rows to convert URLs to hyperlinks
for row in range(2, len(discount_items) + 2):  # Start from the second row
    cell = ws.cell(row=row, column=2)
    url = cell.value
    if url:
        cell.hyperlink = url
        cell.style = "Hyperlink"  # Apply hyperlink style

# Auto-adjust the other columns by content
for col in range(3, len(df.columns) + 1):
    max_length = 0
    column = get_column_letter(col)
    for cell in ws[column]:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the updated Excel file
wb.save(output_file)
